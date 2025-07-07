from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import sys
import time
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, ttk

# Constants - Using absolute path in Documents folder
FILE_PATH = "offline_dtr.xlsx"  # Changed to your requested file path
TIME_FORMAT = "%H:%M:%S"
DATE_FORMAT = "%Y-%m-%d"
MAX_RETRIES = 3
RETRY_DELAY = 1  # seconds

class DTRSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Employee Time In/Out System")
        self.root.geometry("400x300")
        self.root.resizable(False, False)
        
        # Initialize the DTR file
        self.initialize_dtr_file()
        
        # Create UI elements
        self.create_widgets()
        
    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="Employee Time Recording", font=('Helvetica', 14, 'bold'))
        title_label.pack(pady=10)
        
        # Employee ID
        id_frame = ttk.Frame(main_frame)
        id_frame.pack(pady=5, fill=tk.X)
        ttk.Label(id_frame, text="Employee ID:").pack(side=tk.LEFT)
        self.emp_id_entry = ttk.Entry(id_frame)
        self.emp_id_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        
        # Employee Name
        name_frame = ttk.Frame(main_frame)
        name_frame.pack(pady=5, fill=tk.X)
        ttk.Label(name_frame, text="Employee Name:").pack(side=tk.LEFT)
        self.emp_name_entry = ttk.Entry(name_frame)
        self.emp_name_entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
        
        # Action buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        self.time_in_btn = ttk.Button(button_frame, text="Time In", command=lambda: self.record_action("Time In"))
        self.time_in_btn.pack(side=tk.LEFT, padx=10)
        
        self.time_out_btn = ttk.Button(button_frame, text="Time Out", command=lambda: self.record_action("Time Out"))
        self.time_out_btn.pack(side=tk.LEFT, padx=10)
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.pack(fill=tk.X, pady=(20, 0))
        
        # Set focus to ID entry
        self.emp_id_entry.focus_set()
        
    def initialize_dtr_file(self):
        """Initialize the DTR file with proper error handling"""
        try:
            if not os.path.exists(FILE_PATH):
                wb = Workbook()
                ws = wb.active
                ws.title = "DTR Records"
                headers = ["Employee ID", "Name", "Date", "Time", "Action"]
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = cell.font.copy(bold=True)
                ws.freeze_panes = "A2"
                self.safe_save(wb, FILE_PATH)
        except Exception as e:
            messagebox.showerror("Initialization Error", f"Fatal error during initialization: {str(e)}")
            sys.exit(1)

    def safe_save(self, wb, filepath):
        """Save workbook with retries and error handling"""
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                wb.save(filepath)
                return True
            except PermissionError:
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                else:
                    messagebox.showerror("Save Error", 
                        "Max retries reached. Please:\n"
                        "1. Close Excel if the file is open\n"
                        f"2. Check file permissions for: {filepath}")
                    return False
            except Exception as e:
                messagebox.showerror("Save Error", f"Unexpected save error: {str(e)}")
                return False
        return False

    def record_action(self, action):
        """Record the time in/out action"""
        emp_id = self.emp_id_entry.get().strip()
        name = self.emp_name_entry.get().strip()
        
        if not emp_id:
            messagebox.showwarning("Input Error", "Please enter Employee ID")
            self.emp_id_entry.focus_set()
            return
            
        if not name:
            messagebox.showwarning("Input Error", "Please enter Employee Name")
            self.emp_name_entry.focus_set()
            return
            
        try:
            # Get current timestamp
            now = datetime.now()
            date_str = now.strftime(DATE_FORMAT)
            time_str = now.strftime(TIME_FORMAT)
            
            # Load or create workbook
            try:
                wb = load_workbook(FILE_PATH)
            except FileNotFoundError:
                self.initialize_dtr_file()
                wb = load_workbook(FILE_PATH)
                
            ws = wb.active
            
            # Add new record
            ws.append([emp_id, name, date_str, time_str, action])
            
            # Save with error handling
            if self.safe_save(wb, FILE_PATH):
                self.status_var.set(f"Successfully recorded {action} for {name} at {time_str}")
                messagebox.showinfo("Success", f"{action} recorded successfully for {name}")
                self.emp_id_entry.delete(0, tk.END)
                self.emp_name_entry.delete(0, tk.END)
                self.emp_id_entry.focus_set()
            else:
                self.status_var.set(f"Failed to record {action}")
                
        except Exception as e:
            messagebox.showerror("Recording Error", f"Error recording DTR: {str(e)}")
            self.status_var.set("Error occurred")

def main():
    root = tk.Tk()
    app = DTRSystem(root)
    root.mainloop()

if __name__ == "__main__":
    main()